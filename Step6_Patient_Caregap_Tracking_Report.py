#!/usr/bin/env python3
# Step6_Patient_Caregap_Tracking_Report.py
# Updated: 2026-03-31 based on weekly discussion dated 27-03-2026

import argparse
import os
import re
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import Dict, List, Optional, Tuple

import pandas as pd
from azure.identity import DefaultAzureCredential
from azure.storage.blob import BlobServiceClient

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


#STEP4_NAME_TS = re.compile(r"step4_(\d{8})_(\d{6})", re.IGNORECASE)
STEP4_NAME_TS = re.compile(r"WHF_.*_(\d{8})_(\d{6})", re.IGNORECASE)
CPT_5DIGIT = re.compile(r"\b(\d{5})\b")


@dataclass(frozen=True)
class BlobRef:
    container: str
    name: str


def _log(msg: str) -> None:
    print(msg)


def _normalize_patient_id(x) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    s = str(x).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def _normalize_cpt(x) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    s = str(x).strip()
    m = CPT_5DIGIT.search(s)
    return m.group(1) if m else s


def _fmt_date(x) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)) or pd.isna(x):
        return ""
    try:
        return pd.to_datetime(x).strftime("%Y-%m-%d")
    except Exception:
        return ""


def _parse_step4_ts(blob_name: str) -> Optional[datetime]:
    m = STEP4_NAME_TS.search(blob_name or "")
    if not m:
        return None
    ymd, hms = m.group(1), m.group(2)
    try:
        return datetime.strptime(f"{ymd}_{hms}", "%Y%m%d_%H%M%S")
    except Exception:
        return None


def _conn_string_looks_valid(conn: str) -> bool:
    c = (conn or "").strip()
    if not c:
        return False
    return ("DefaultEndpointsProtocol=" in c) and ("AccountName=" in c) and ("AccountKey=" in c)


def get_blob_service_client(account_url: str) -> BlobServiceClient:
    conn = os.getenv("AZURE_STORAGE_CONNECTION_STRING", "")
    if _conn_string_looks_valid(conn):
        try:
            return BlobServiceClient.from_connection_string(conn.strip())
        except Exception:
            pass

    return BlobServiceClient(
        account_url=account_url,
        credential=DefaultAzureCredential(exclude_interactive_browser_credential=True),
    )


def download_bytes(bsc: BlobServiceClient, ref: BlobRef) -> bytes:
    bc = bsc.get_blob_client(container=ref.container, blob=ref.name)
    return bc.download_blob().readall()


def upload_bytes(bsc: BlobServiceClient, ref: BlobRef, data: bytes) -> None:
    bc = bsc.get_blob_client(container=ref.container, blob=ref.name)
    bc.upload_blob(data, overwrite=True)


def pick_latest_step4(
    bsc: BlobServiceClient, container: str, prefix: str) -> Tuple[str, Optional[datetime]]:
    
    prefix = (prefix or "").lstrip("/")
    if prefix and not prefix.endswith("/"):
        prefix += "/"

    cc = bsc.get_container_client(container)
    found: List[Tuple[Optional[datetime], datetime, str]] = []

    for b in cc.list_blobs(name_starts_with=prefix + "WHF_"):
        name = b.name
        low = name.lower()

       # if "care_tracking_output" not in low:
        if "current_pregnant_patients" not in low:
            continue
        if not (low.endswith(".csv") or low.endswith(".csv.csv")):
            continue

        ts = _parse_step4_ts(name)
        lm = b.last_modified
        if lm and lm.tzinfo is not None:
            lm = lm.replace(tzinfo=None)

        found.append((ts, lm or datetime.min, name))

    if not found:
        raise FileNotFoundError(f"No Step4 outputs found in '{container}'.")

    #Pick latest only
    found.sort(key=lambda x: (x[0] or datetime.min, x[1]))
    latest_ts, _, latest_name = found[-1]

    return latest_name, latest_ts


def load_step4_patients(raw_csv: bytes) -> pd.DataFrame:
    df = pd.read_csv(BytesIO(raw_csv), dtype=str, keep_default_na=False)

    if "patient_id" not in df.columns:
        for cand in ["patientid", "PatientID", "PATIENT_ID", "MRN"]:
            if cand in df.columns:
                df = df.rename(columns={cand: "patient_id"})
                break

    if "patient_id" not in df.columns:
        raise KeyError(f"Step4 file missing patient_id. Columns found: {list(df.columns)}")

    df["patient_id"] = df["patient_id"].apply(_normalize_patient_id)
    return df


def auto_width(ws) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)


def write_report(rows: List[Dict], meta: Dict, out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pregnant Patients for Followup"
    ws["A1"] = "Pregnant Patients for Followup Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Generated"
    ws["B2"] = meta.get("generated_ts", "")    
    ws["A3"] = "Current Step4"
    ws["B3"] = meta.get("latest_step4_blob", "")    
    ws["A4"] = "Patient count 1 week or more followup gap for gestational age between 35 and 41"
    ws["B4"] = meta.get("ga_35_count", 0)
    ws["A5"] = "Patient count for 4 weeks or more followup gap"
    ws["B5"] = meta.get("gap_4week_count", 0)

    header_row = 6
    headers = [
        #"priority",        
        "patient_id",
        "last_encounter_date",
        "gestational age",
        "notes",
    ]

    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=j, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for i, r in enumerate(rows, header_row + 1):
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=i, column=j, value=r.get(h, ""))
            c.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = ws["A6"]

    last_row = header_row + len(rows)
    if last_row >= header_row + 1:
        tbl = Table(displayName="Followup", ref=f"A{header_row}:I{last_row}")
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(tbl)

    auto_width(ws)
    
    wb.save(out_path)
    
def main() -> None:
    ap = argparse.ArgumentParser(description="Create PregnantPatientsforFollowup_<date>.xlsx from Step4.")
    ap.add_argument("--account-name", required=True)

    ap.add_argument("--step4-container", default="output")
    ap.add_argument("--step4-prefix", default="")

   #ap.add_argument("--encounters-container", default="input")
   #ap.add_argument("--encounters-blob", default="EncounterData.xlsx")

    ap.add_argument("--output-container", default="output")
    ap.add_argument("--output-prefix", default="")

    #ap.add_argument("--grace-days", type=int, default=7)
    ap.add_argument("--local-out", default="")
    args = ap.parse_args()

    account_url = f"https://{args.account_name}.blob.core.windows.net"
    bsc = get_blob_service_client(account_url)

    
    latest_blob,latest_ts = pick_latest_step4(bsc, args.step4_container, args.step4_prefix)    
    
    _log(f"[CAREGAP] Current  Step4: {latest_blob}")

    
    latest_df = load_step4_patients(download_bytes(bsc, BlobRef(args.step4_container, latest_blob)))
    
    # filter patients with GA > 35 and latest_encounter_date - current_date >=7 days
    
    # Ensure GA is numeric
    latest_df["current_gestational_age"] = pd.to_numeric(latest_df["current_gestational_age"], errors="coerce")
    
    # Ensure datetime conversion
    latest_df["latest_encounter_date"] = pd.to_datetime(latest_df["latest_encounter_date"], errors="coerce")
    
    # Get current date (normalized to remove time component if needed)
    current_date = pd.Timestamp.today().normalize()
    
    # Apply both conditions( ga> 35 and current_date- last_encounter_date >=7 
    high_ga_35_df = latest_df[(latest_df["current_gestational_age"] > 35) & (latest_df["current_gestational_age"] <= 41) 
    &((current_date - latest_df["latest_encounter_date"]).dt.days >= 7)].copy()
    
    ga_35_count = len(high_ga_35_df)
    
    # Apply both conditions  current_date- last_encounter_date >=28 
    gap_4week_df = latest_df[(current_date - latest_df["latest_encounter_date"]).dt.days >= 28].copy()
    gap_4week_count = len(gap_4week_df)
    
    rows: List[Dict] = []    
    for _, row in high_ga_35_df.iterrows():
        rows.append({
            "priority": 1,
            "patient_id": row["patient_id"],
            "last_encounter_date": row["latest_encounter_date"].strftime("%m/%d/%Y"),
            "gestational age": row["current_gestational_age"],
            "notes": "1 week or more followup gap for gestational age between 35 and 41",
    })
        
        
    existing_pids = set(r["patient_id"] for r in rows)
    for _, row in gap_4week_df.iterrows():
        if row["patient_id"] not in existing_pids:
            rows.append({
            "priority": 2,
            "patient_id": row["patient_id"],
            "last_encounter_date": row["latest_encounter_date"].strftime("%m/%d/%Y"),
            "gestational age": row["current_gestational_age"],
            "notes": "4 weeks or more followup gap",
            })

    #rows.sort(key=lambda r: (r.get("priority", 9), str(r.get("patient_id", ""))))
    # Sort by gestational age DESCENDING added 14-04-2026
    rows.sort(key=lambda r: r.get("current_gestational_age", ""), reverse=True)

    report_date = datetime.now().strftime("%Y%m%d")
    
    #local_out = args.local_out.strip() or f"CaregapTrackingReport_{report_date}.xlsx" 
    report_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    local_out = args.local_out.strip() or f"/tmp/Pregnant_Patients_for_Followup_{report_ts}.xlsx"

    write_report(
        rows=rows,
        #evidence_df=delivery_win,
        meta={
            "generated_ts": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
            #"prev_step4_blob": prev_blob,
            "latest_step4_blob": latest_blob,
            #"window_start": window_start_print,
            #"window_end": window_end_print,
            #"dropped_count": len(dropped),
            "ga_35_count":ga_35_count,
            "gap_4week_count":gap_4week_count
        },
        out_path=local_out,
    )

    print(f"[DEBUG] File path: {local_out}")
    print(f"[DEBUG] Exists? {os.path.exists(local_out)}")
    
    _log(f"[CAREGAP] Wrote local: {local_out} ({len(rows)}  patients with caregap)")

    out_prefix = (args.output_prefix or "").lstrip("/")
    if out_prefix and not out_prefix.endswith("/"):
        out_prefix += "/"
    out_blob = f"{out_prefix}{os.path.basename(local_out)}"
    
    if not os.path.exists(local_out):
        raise FileNotFoundError(f"File not created: {local_out}")


    with open(local_out, "rb") as f:
        upload_bytes(bsc, BlobRef(args.output_container, out_blob), f.read())

    _log(f"[CAREGAP] Uploaded: container={args.output_container} blob={out_blob}")


if __name__ == "__main__":
    main()
