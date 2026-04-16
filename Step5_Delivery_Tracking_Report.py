#!/usr/bin/env python3
# Step5_Delivery_Tracking_Report_update_clean.py
# Updated: 2026-02-21

import argparse
import os
import re
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import Dict, List, Optional, Tuple

import pandas as pd
from azure.identity import DefaultAzureCredential
from azure.storage.blob import BlobServiceClient

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


DELIVERY_CPTS = {
    "59400", "59409", "59410", "59412", "59414",
    "59510", "59514", "59515",
    "59610", "59612", "59614", "59618", "59620", "59622",
}

#added on 14-04-2026
#STEP4_NAME_TS = re.compile(r"step4_(\d{8})_(\d{6})", re.IGNORECASE)
STEP4_NAME_TS = re.compile(r"WHF_.*_(\d{8})_(\d{6})", re.IGNORECASE)
CPT_5DIGIT = re.compile(r"\b(\d{5})\b")


@dataclass(frozen=True)
class BlobRef:
    container: str
    name: str


def _log(msg: str) -> None:
    print(msg)


def _normalize_patient_id(x) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    s = str(x).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def _normalize_cpt(x) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    s = str(x).strip()
    m = CPT_5DIGIT.search(s)
    return m.group(1) if m else s


def _fmt_date(x) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)) or pd.isna(x):
        return ""
    try:
        return pd.to_datetime(x).strftime("%Y-%m-%d")
    except Exception:
        return ""


def _parse_step4_ts(blob_name: str) -> Optional[datetime]:
    m = STEP4_NAME_TS.search(blob_name or "")
    if not m:
        return None
    ymd, hms = m.group(1), m.group(2)
    try:
        return datetime.strptime(f"{ymd}_{hms}", "%Y%m%d_%H%M%S")
    except Exception:
        return None


def _conn_string_looks_valid(conn: str) -> bool:
    c = (conn or "").strip()
    if not c:
        return False
    return ("DefaultEndpointsProtocol=" in c) and ("AccountName=" in c) and ("AccountKey=" in c)


def get_blob_service_client(account_url: str) -> BlobServiceClient:
    conn = os.getenv("AZURE_STORAGE_CONNECTION_STRING", "")
    if _conn_string_looks_valid(conn):
        try:
            return BlobServiceClient.from_connection_string(conn.strip())
        except Exception:
            pass

    return BlobServiceClient(
        account_url=account_url,
        credential=DefaultAzureCredential(exclude_interactive_browser_credential=True),
    )


def download_bytes(bsc: BlobServiceClient, ref: BlobRef) -> bytes:
    bc = bsc.get_blob_client(container=ref.container, blob=ref.name)
    return bc.download_blob().readall()


def upload_bytes(bsc: BlobServiceClient, ref: BlobRef, data: bytes) -> None:
    bc = bsc.get_blob_client(container=ref.container, blob=ref.name)
    bc.upload_blob(data, overwrite=True)


def pick_latest_two_step4(
    bsc: BlobServiceClient, container: str, prefix: str
) -> Tuple[str, str, Optional[datetime], Optional[datetime]]:
    prefix = (prefix or "").lstrip("/")
    if prefix and not prefix.endswith("/"):
        prefix += "/"

    cc = bsc.get_container_client(container)
    found: List[Tuple[Optional[datetime], datetime, str]] = []

    #for b in cc.list_blobs(name_starts_with=prefix + "step4_"):
    for b in cc.list_blobs(name_starts_with=prefix + "WHF_"):
        name = b.name
        low = name.lower()
        #added on 14-04-2026
        #if "care_tracking_output" not in low:
        if "current_pregnant_patients" not in low:
            continue
        if not (low.endswith(".csv") or low.endswith(".csv.csv")):
            continue

        ts = _parse_step4_ts(name)
        lm = b.last_modified
        if lm and lm.tzinfo is not None:
            lm = lm.replace(tzinfo=None)

        found.append((ts, lm or datetime.min, name))

    if len(found) < 2:
        raise FileNotFoundError(f"Need at least 2 Step4 outputs to compare. Found {len(found)} in '{container}'.")

    found.sort(key=lambda x: (x[0] or datetime.min, x[1]))
    prev_ts, _, prev_name = found[-2]
    latest_ts, _, latest_name = found[-1]
    return prev_name, latest_name, prev_ts, latest_ts


def load_step4_patients(raw_csv: bytes) -> pd.DataFrame:
    df = pd.read_csv(BytesIO(raw_csv), dtype=str, keep_default_na=False)

    if "patient_id" not in df.columns:
        for cand in ["patientid", "PatientID", "PATIENT_ID", "MRN"]:
            if cand in df.columns:
                df = df.rename(columns={cand: "patient_id"})
                break

    if "patient_id" not in df.columns:
        raise KeyError(f"Step4 file missing patient_id. Columns found: {list(df.columns)}")

    df["patient_id"] = df["patient_id"].apply(_normalize_patient_id)
    return df


def load_encounters(raw_xlsx: bytes) -> pd.DataFrame:
    enc = pd.read_excel(BytesIO(raw_xlsx))

    def find_col(cands: List[str]) -> str:
        lower = {c.lower(): c for c in enc.columns}
        for c in cands:
            if c in enc.columns:
                return c
            if c.lower() in lower:
                return lower[c.lower()]
        raise KeyError(f"Missing column. Tried {cands}. Found {list(enc.columns)}")

    pid_col = find_col(["patientid", "patient_id", "PatientID", "PATIENT_ID", "MRN"])
    cpt_col = find_col(["CPTCode", "CPT_CODE", "CPT", "cpt"])
    dos_col = find_col(["DOS", "Date of Service", "Service Date"])

    desc_col = None
    lowered = {c.lower(): c for c in enc.columns}
    for cand in ["CPTDescription", "cpt_description", "ProcedureDescription", "Procedure Desc", "CPT Desc"]:
        if cand in enc.columns:
            desc_col = cand
            break
        if cand.lower() in lowered:
            desc_col = lowered[cand.lower()]
            break

    enc = enc.rename(columns={pid_col: "patient_id", cpt_col: "cpt", dos_col: "dos"})
    enc["patient_id"] = enc["patient_id"].apply(_normalize_patient_id)
    enc["cpt"] = enc["cpt"].apply(_normalize_cpt)
    enc["dos"] = pd.to_datetime(enc["dos"], errors="coerce")

    if desc_col:
        enc = enc.rename(columns={desc_col: "cpt_description"})
    else:
        enc["cpt_description"] = ""

    return enc[["patient_id", "dos", "cpt", "cpt_description"]].copy()


def auto_width(ws) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)


def write_report(rows: List[Dict], evidence_df: pd.DataFrame, meta: Dict, out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery Tracking"

    ws["A1"] = "Delivery Tracking Report"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = "Generated"
    ws["B2"] = meta.get("generated_ts", "")
    ws["A3"] = "Previous Step4"
    ws["B3"] = meta.get("prev_step4_blob", "")
    ws["A4"] = "Current Step4"
    ws["B4"] = meta.get("latest_step4_blob", "")
    ws["A5"] = "Drop window start"
    ws["B5"] = meta.get("window_start", "")
    ws["A6"] = "Drop window end"
    ws["B6"] = meta.get("window_end", "")
    ws["A7"] = "Count of Dropped patients"
    ws["B7"] = meta.get("dropped_count", 0)
    ws["A8"] = "Count of patients with Gestational age > 41 "
    ws["B8"] = meta.get("ga_41_count", 0)

    header_row = 9
    headers = [
        "priority",
        "patient_id",
        "first_encounter_date",
        "start_of_pregnancy_date",
        "delivery_in_window",
        "delivery_date_in_window",
        "delivery_cpt_in_window",
        "delivery_anytime",
        "delivery_date_anytime",
        "delivery_cpt_anytime",
        "notes",
    ]

    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=j, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for i, r in enumerate(rows, header_row + 1):
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=i, column=j, value=r.get(h, ""))
            c.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = ws["A10"]

    last_row = header_row + len(rows)
    if last_row >= header_row + 1:
        tbl = Table(displayName="DeliveryTracking", ref=f"A{header_row}:I{last_row}")
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(tbl)

    auto_width(ws)

    ws2 = wb.create_sheet("Encounter Evidence")
    ev_cols = ["patient_id", "dos", "cpt", "cpt_description"]

    for j, h in enumerate(ev_cols, 1):
        ws2.cell(row=1, column=j, value=h).font = Font(bold=True)

    if not evidence_df.empty:
        ev = evidence_df.copy()
        ev["dos"] = ev["dos"].apply(_fmt_date)
        for i, rec in enumerate(ev[ev_cols].to_dict(orient="records"), 2):
            for j, h in enumerate(ev_cols, 1):
                ws2.cell(row=i, column=j, value=rec.get(h, ""))

    ws2.freeze_panes = ws2["A2"]
    auto_width(ws2)

    wb.save(out_path)


def main() -> None:
    ap = argparse.ArgumentParser(description="Create DeliveryTrackingReport_<date>.xlsx from Step4 drop-offs.")
    ap.add_argument("--account-name", required=True)

    ap.add_argument("--step4-container", default="output")
    ap.add_argument("--step4-prefix", default="")

    ap.add_argument("--encounters-container", default="input")
    ap.add_argument("--encounters-blob", default="EncounterData.xlsx")

    ap.add_argument("--output-container", default="output")
    ap.add_argument("--output-prefix", default="")

    ap.add_argument("--grace-days", type=int, default=7)
    ap.add_argument("--local-out", default="")
    args = ap.parse_args()

    account_url = f"https://{args.account_name}.blob.core.windows.net"
    bsc = get_blob_service_client(account_url)

    prev_blob, latest_blob, prev_ts, latest_ts = pick_latest_two_step4(bsc, args.step4_container, args.step4_prefix)
    _log(f"[DELIVERY] Previous Step4: {prev_blob}")
    _log(f"[DELIVERY] Current  Step4: {latest_blob}")

    prev_df = load_step4_patients(download_bytes(bsc, BlobRef(args.step4_container, prev_blob)))
    latest_df = load_step4_patients(download_bytes(bsc, BlobRef(args.step4_container, latest_blob)))

    dropped = sorted(set(prev_df["patient_id"]) - set(latest_df["patient_id"]))
    dropped_set = set(dropped)

    enc_df = load_encounters(download_bytes(bsc, BlobRef(args.encounters_container, args.encounters_blob)))
    
    # added as per discussion on 27-03-2026
    # filter patients with GA > 41
    latest_df["current_gestational_age"] = pd.to_numeric(latest_df["current_gestational_age"], errors="coerce")
    high_ga_df = latest_df[latest_df["current_gestational_age"] > 41].copy()
    
    ga_41_count = len(high_ga_df)

    delivery_all = enc_df[(enc_df["patient_id"].isin(dropped_set)) & (enc_df["cpt"].isin(DELIVERY_CPTS))].copy()
    delivery_all = delivery_all.dropna(subset=["dos"]).sort_values(["patient_id", "dos", "cpt"])

    window_start = prev_ts.date() if prev_ts else None
    window_end = latest_ts.date() if latest_ts else None

    if window_start and window_end:
        ws_dt = datetime(window_start.year, window_start.month, window_start.day)
        we_dt = datetime(window_end.year, window_end.month, window_end.day) + timedelta(days=args.grace_days)
        delivery_win = delivery_all[(delivery_all["dos"] >= ws_dt) & (delivery_all["dos"] <= we_dt)].copy()
        window_start_print = window_start.isoformat()
        window_end_print = we_dt.date().isoformat()
    else:
        delivery_win = delivery_all.copy()
        window_start_print = "unknown"
        window_end_print = "unknown"

    by_pid_any = {pid: g.copy() for pid, g in delivery_all.groupby("patient_id")} if not delivery_all.empty else {}
    by_pid_win = {pid: g.copy() for pid, g in delivery_win.groupby("patient_id")} if not delivery_win.empty else {}

    def last_date_and_codes(g: Optional[pd.DataFrame]) -> Tuple[str, str]:
        if g is None or g.empty:
            return "", ""
        return _fmt_date(g["dos"].max()), ", ".join(sorted(set(g["cpt"].tolist())))

    rows: List[Dict] = []
    
    # consider previous lookup as latest lookup no data for corresponding patient
    prev_lookup = prev_df.set_index("patient_id").to_dict(orient="index")
    for pid in dropped:
        g_any = by_pid_any.get(pid)
        g_win = by_pid_win.get(pid)

        any_found = g_any is not None and not g_any.empty
        win_found = g_win is not None and not g_win.empty

        if not any_found:
            priority = 1
            note = "No delivery CPT found anywhere in EncounterData. Needs review."
        elif not win_found:
            priority = 2
            note = "Delivery CPT exists, but not inside the expected drop-off window. Review timing."
        else:
            priority = 3
            note = "Delivery CPT found in expected window."
        

        win_date, win_codes = last_date_and_codes(g_win)
        any_date, any_codes = last_date_and_codes(g_any)   

        patient_info = prev_lookup.get(pid, {})
    

        rows.append({
            "priority": priority,
            "patient_id": pid,
            "first_encounter_date" :patient_info.get("first_encounter_date", ""),
            "start_of_pregnancy_date" :patient_info.get("start_of_pregnancy_date", ""),
            "delivery_in_window": "Y" if win_found else "N",
            "delivery_date_in_window": win_date,
            "delivery_cpt_in_window": win_codes,
            "delivery_anytime": "Y" if any_found else "N",
            "delivery_date_anytime": any_date,
            "delivery_cpt_anytime": any_codes,
            "notes": note,
        })
        
        # added as per discussion on 27-03-2026
        
        existing_pids = set(r["patient_id"] for r in rows)
        for pid in high_ga_df["patient_id"].unique():
            if pid not in existing_pids:
                rows.append({
                    "priority": 4,
                    "patient_id": pid,
                    "first_encounter_date" : "",
                    "start_of_pregnancy_date" : "",
                    "delivery_in_window": "N",
                    "delivery_date_in_window": "",
                    "delivery_cpt_in_window": "",
                    "delivery_anytime": "N",
                    "delivery_date_anytime": "",
                    "delivery_cpt_anytime": "",
                    "notes": "Gestational Age > 41 Weeks."
                })

    rows.sort(key=lambda r: (r.get("priority", 9), str(r.get("patient_id", ""))))

    #added 14-04-2026
    #report_date = datetime.now().strftime("%Y%m%d")
    report_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    local_out = args.local_out.strip() or f"Delivery_Tracking_Report_{report_ts}.xlsx"

    write_report(
        rows=rows,
        evidence_df=delivery_win,
        meta={
            "generated_ts": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
            "prev_step4_blob": prev_blob,
            "latest_step4_blob": latest_blob,
            "window_start": window_start_print,
            "window_end": window_end_print,
            "dropped_count": len(dropped),
            "ga_41_count":ga_41_count,
        },
        out_path=local_out,
    )

    _log(f"[DELIVERY] Wrote local: {local_out} ({len(rows)} dropped patients)")

    out_prefix = (args.output_prefix or "").lstrip("/")
    if out_prefix and not out_prefix.endswith("/"):
        out_prefix += "/"
    out_blob = f"{out_prefix}{os.path.basename(local_out)}"

    with open(local_out, "rb") as f:
        upload_bytes(bsc, BlobRef(args.output_container, out_blob), f.read())

    _log(f"[DELIVERY] Uploaded: container={args.output_container} blob={out_blob}")


if __name__ == "__main__":
    main()
