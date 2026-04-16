#!/usr/bin/env python3
# Step7_Hideen_Pregnant_Patient_Report.py
# Updated: 2026-04-12

import argparse
import os
import re
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import Dict, List, Optional, Tuple

import pandas as pd
from azure.identity import DefaultAzureCredential
from azure.storage.blob import BlobServiceClient

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


HOSPITAL_LOCATIONS = {
"WHF - SAMC - ER", 
"WHF - SAMC - IP",
"WHF - SAMC - OP",
"WHF - ABMC - IP",
"WHF - HOFFMAN ESTATES SURGERY CENTER, LLC",
"WHF - NW COMMUNITY - ER",
"WHF - NW COMMUNITY - IP",
"WHF - NW COMMUNITY - OP",
}

#STEP4_NAME_TS = re.compile(r"step4_(\d{8})_(\d{6})", re.IGNORECASE)
STEP4_NAME_TS = re.compile(r"WHF_(\d{8})_(\d{6})", re.IGNORECASE)
CPT_5DIGIT = re.compile(r"\b(\d{5})\b")


@dataclass(frozen=True)
class BlobRef:
    container: str
    name: str


def _log(msg: str) -> None:
    print(msg)


def _normalize_patient_id(x) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    s = str(x).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def _normalize_cpt(x) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    s = str(x).strip()
    m = CPT_5DIGIT.search(s)
    return m.group(1) if m else s


def _fmt_date(x) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)) or pd.isna(x):
        return ""
    try:
        return pd.to_datetime(x).strftime("%Y-%m-%d")
    except Exception:
        return ""


def _parse_step4_ts(blob_name: str) -> Optional[datetime]:
    m = STEP4_NAME_TS.search(blob_name or "")
    if not m:
        return None
    ymd, hms = m.group(1), m.group(2)
    try:
        return datetime.strptime(f"{ymd}_{hms}", "%Y%m%d_%H%M%S")
    except Exception:
        return None


def _conn_string_looks_valid(conn: str) -> bool:
    c = (conn or "").strip()
    if not c:
        return False
    return ("DefaultEndpointsProtocol=" in c) and ("AccountName=" in c) and ("AccountKey=" in c)


def get_blob_service_client(account_url: str) -> BlobServiceClient:
    conn = os.getenv("AZURE_STORAGE_CONNECTION_STRING", "")
    if _conn_string_looks_valid(conn):
        try:
            return BlobServiceClient.from_connection_string(conn.strip())
        except Exception:
            pass

    return BlobServiceClient(
        account_url=account_url,
        credential=DefaultAzureCredential(exclude_interactive_browser_credential=True),
    )


def download_bytes(bsc: BlobServiceClient, ref: BlobRef) -> bytes:
    bc = bsc.get_blob_client(container=ref.container, blob=ref.name)
    return bc.download_blob().readall()


def upload_bytes(bsc: BlobServiceClient, ref: BlobRef, data: bytes) -> None:
    bc = bsc.get_blob_client(container=ref.container, blob=ref.name)
    bc.upload_blob(data, overwrite=True)

#Pickup latest step4 file
def pick_latest_step4(
    bsc: BlobServiceClient, container: str, prefix: str) -> Tuple[str, Optional[datetime]]:
    
    prefix = (prefix or "").lstrip("/")
    if prefix and not prefix.endswith("/"):
        prefix += "/"

    cc = bsc.get_container_client(container)
    found: List[Tuple[Optional[datetime], datetime, str]] = []

    for b in cc.list_blobs(name_starts_with=prefix + "WHF_"):
        name = b.name
        low = name.lower()

        #if "care_tracking_output" not in low:
        if "current_pregnant_patients" not in low:
            continue
        if not (low.endswith(".csv") or low.endswith(".csv.csv")):
            continue

        ts = _parse_step4_ts(name)
        lm = b.last_modified
        if lm and lm.tzinfo is not None:
            lm = lm.replace(tzinfo=None)

        found.append((ts, lm or datetime.min, name))

    if not found:
        raise FileNotFoundError(f"No Step4 outputs found in '{container}'.")

    #Pick latest only
    found.sort(key=lambda x: (x[0] or datetime.min, x[1]))
    latest_ts, _, latest_name = found[-1]

    return latest_name, latest_ts

def load_step4_patients(raw_csv: bytes) -> pd.DataFrame:
    df = pd.read_csv(BytesIO(raw_csv), dtype=str, keep_default_na=False)

    if "patient_id" not in df.columns:
        for cand in ["patientid", "PatientID", "PATIENT_ID", "MRN"]:
            if cand in df.columns:
                df = df.rename(columns={cand: "patient_id"})
                break

    if "patient_id" not in df.columns:
        raise KeyError(f"Step4 file missing patient_id. Columns found: {list(df.columns)}")

    df["patient_id"] = df["patient_id"].apply(_normalize_patient_id)
    return df


def load_encounters(raw_xlsx: bytes) -> pd.DataFrame:
    enc = pd.read_excel(BytesIO(raw_xlsx), parse_dates=["DOS", "DOB"])
    
    def find_col(cands: List[str]) -> str:       
        
        lower = {c.lower(): c for c in enc.columns}
        for c in cands:
            if c in enc.columns:
                return c
            if c.lower() in lower:
                return lower[c.lower()]
        raise KeyError(f"Missing column. Tried {cands}. Found {list(enc.columns)}")

    pid_col = find_col(["patientid", "patient_id", "PatientID", "PATIENT_ID", "MRN"])
    dob_col = find_col(["DOB", "DateOfBirth", "Birth Date", "dob"])
    cpt_col = find_col(["CPTCode", "CPT_CODE", "CPT", "cpt"])
    dos_col = find_col(["DOS", "Date of Service", "Service Date"])

    desc_col = None
    lowered = {c.lower(): c for c in enc.columns}
    for cand in ["CPTDescription", "cpt_description", "ProcedureDescription", "Procedure Desc", "CPT Desc"]:
        if cand in enc.columns:
            desc_col = cand
            break
        if cand.lower() in lowered:
            desc_col = lowered[cand.lower()]
            break

    #enc = enc.rename(columns={pid_col: "patient_id", cpt_col: "cpt", dos_col: "dos"})

    #enc = enc.rename(columns={pid_col: "patient_id", cpt_col: "cpt", dos_col: "dos"})
    enc = enc.rename(
            columns={
                pid_col:            "patient_id",
                "patient lastname":     "last_name",
                "patient firstname":    "first_name",
                dob_col:                  "date_of_birth",
                dos_col:                  "encounter_date",
                "ServiceDepartment":    "location",
                cpt_col:              "cpt_code",
                "DiagICD_10_1":         "ICD_10_1",
                "DiagICD_10_2":         "ICD_10_2",
                "DiagICD_10_3":         "ICD_10_3",
                "DiagICD_10_4":         "ICD_10_4",
            }
        )
    
    enc["patient_id"] = enc["patient_id"].apply(_normalize_patient_id)
    enc["cpt_code"] = enc["cpt_code"].apply(_normalize_cpt)
    enc["encounter_date"] = pd.to_datetime(enc["encounter_date"], errors="coerce")
    enc["date_of_birth"] = pd.to_datetime(enc["date_of_birth"], errors="coerce")


    return enc[["patient_id", "first_name", "last_name","date_of_birth", "encounter_date", "location", "cpt_code", "ICD_10_1","ICD_10_2","ICD_10_3","ICD_10_4"]].copy()


def auto_width(ws) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)


def write_report(rows: List[Dict], evidence_df: pd.DataFrame, meta: Dict, out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "FQHC_visits "

    ws["A1"] = "FQHC_visits  Report"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = "Generated"
    ws["B2"] = meta.get("generated_ts", "")
    
    #ws["A3"] = "Previous Step4"
    #ws["B3"] = meta.get("prev_step4_blob", "")
    
    ws["A4"] = "Current Step4"
    ws["B4"] = meta.get("latest_step4_blob", "")
    
    #ws["A5"] = "Drop window start"
    #ws["B5"] = meta.get("window_start", "")
    
    #ws["A6"] = "Drop window end"
    #ws["B6"] = meta.get("window_end", "")
    
    ws["A7"] = "FQHC_visits  Patients"
    ws["B7"] = meta.get("hidden_count", 0)
    

    header_row = 9
    headers = [
        "patient_id", 
        "last_name", 
        "first_name", 
        "date_of_birth",
        "encounter_date", 
        "location", 
        "cpt_code", 
        "ICD_10_1",
        "ICD_10_2",
        "ICD_10_3",
        "ICD_10_4",
    ]

    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=j, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for i, r in enumerate(rows, header_row + 1):
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=i, column=j, value=r.get(h, ""))
            c.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = ws["A9"]

    last_row = header_row + len(rows)
    if last_row >= header_row + 1:
        tbl = Table(displayName="FQHC", ref=f"A{header_row}:M{last_row}")
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(tbl)

    auto_width(ws)
    wb.save(out_path)


def main() -> None:
    ap = argparse.ArgumentParser(description="Create FQHC_visits _<date>.xlsx from Step4  and encounter file")
    ap.add_argument("--account-name", required=True)

    ap.add_argument("--step4-container", default="output")
    ap.add_argument("--step4-prefix", default="")

    ap.add_argument("--encounters-container", default="input")
    ap.add_argument("--encounters-blob", default="EncounterData.xlsx")

    ap.add_argument("--output-container", default="output")
    ap.add_argument("--output-prefix", default="")

    #ap.add_argument("--grace-days", type=int, default=7)
    
    ap.add_argument("--local-out", default="")
    args = ap.parse_args()

    account_url = f"https://{args.account_name}.blob.core.windows.net"
    bsc = get_blob_service_client(account_url)

    #prev_blob, latest_blob, prev_ts, latest_ts = pick_latest_two_step4(bsc, args.step4_container, args.step4_prefix)
    latest_blob,latest_ts = pick_latest_step4(bsc, args.step4_container, args.step4_prefix)   
    #_log(f"[DELIVERY] Previous Step4: {prev_blob}")
    _log(f"[FQHC] Current  Step4: {latest_blob}")

    #prev_df = load_step4_patients(download_bytes(bsc, BlobRef(args.step4_container, prev_blob)))
    
    latest_df = load_step4_patients(download_bytes(bsc, BlobRef(args.step4_container, latest_blob)))    
    enc_df = load_encounters(download_bytes(bsc, BlobRef(args.encounters_container, args.encounters_blob)))
    
    diag_cols = ['ICD_10_1','ICD_10_2','ICD_10_3','ICD_10_4']

    diag_mask = (
        enc_df[diag_cols]
        .fillna("")
        .apply(lambda col: col.str.strip().str.upper().str.match(r"^O\d", na=False))
        .any(axis=1)
    )

    mask = (
        enc_df["location"].isin(HOSPITAL_LOCATIONS)
        & enc_df["cpt_code"].astype(str).str.match(r"^99\d{3}")
        & diag_mask
    )

    enc_df_hd_prg = enc_df[mask].copy()

    # Remove already known pregnant patients
    latest_set = set(latest_df["patient_id"])

    hidden_df = enc_df_hd_prg[~enc_df_hd_prg["patient_id"].isin(latest_set)].copy()

    #hidden_df = hidden_df.dropna(subset=["encounter_date"]) \.sort_values(["patient_id", "encounter_date"])
    hidden_df = (hidden_df.dropna(subset=["encounter_date"]).sort_values(by="encounter_date", ascending=False))
    
    hidden_df = hidden_df.drop_duplicates(subset=["patient_id"], keep="first")
    
    hidden_count = hidden_df["patient_id"].nunique()

    rows = []
    for _, row in hidden_df.iterrows():
            rows.append({
                "patient_id": row["patient_id"],
                "last_name": row["last_name"],
                "first_name": row["first_name"],
                #"date_of_birth": _fmt_date(row["date_of_birth"].strftime("%Y-%m-%d %H:%M:%S UTC")),
                "date_of_birth": _fmt_date(row["date_of_birth"]),
                #"encounter_date": _fmt_date(row["encounter_date"].strftime("%Y-%m-%d %H:%M:%S UTC")),
                "encounter_date": _fmt_date(row["encounter_date"]),
                "location": row["location"],
                "cpt_code": row["cpt_code"],
                "ICD_10_1": row.get("ICD_10_1", ""),
                "ICD_10_2": row.get("ICD_10_2", ""),
                "ICD_10_3": row.get("ICD_10_3", ""),
                "ICD_10_4": row.get("ICD_10_4", ""),
            })

    #rows.sort(key=lambda r: r["encounter_date"], reverse=True
   
    report_date = datetime.now().strftime("%Y%m%d")
    local_out = args.local_out.strip() or f"FQHC_visits_{report_date}.xlsx"

    write_report(
        rows=rows,
        evidence_df=None,
        meta={
            "generated_ts": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
            #"prev_step4_blob": prev_blob,
            "latest_step4_blob": latest_blob,
            #"window_start": window_start_print,
            #"window_end": window_end_print,
            "hidden_count": hidden_count,            
        },
        out_path=local_out,
    )

    _log(f"[FQHC] Wrote local: {local_out} ({len(rows)} dropped patients)")

    out_prefix = (args.output_prefix or "").lstrip("/")
    if out_prefix and not out_prefix.endswith("/"):
        out_prefix += "/"
    out_blob = f"{out_prefix}{os.path.basename(local_out)}"

    with open(local_out, "rb") as f:
        upload_bytes(bsc, BlobRef(args.output_container, out_blob), f.read())

    _log(f"[FQHC] Uploaded: container={args.output_container} blob={out_blob}")


if __name__ == "__main__":
    main()
